# =============================
# ORGANON DATA SOLUTIONS
# Fonctions utilitaires
# =============================

import os
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle, PageBreak
from docx import Document
from docx.shared import Inches, Pt, RGBColor
import openpyxl
from config import COLORS, LAYOUT, PATHS, FONTS, ORGANON_SIGNATURE, PALETTE


# =============================
# 1. CHARGEMENT DES DONNÉES
# =============================

def charger_csv(nom_fichier):
    """
    Charge un CSV depuis le dossier data/
    Usage : df = charger_csv("projet1.csv")
    """
    chemin = os.path.join(PATHS["data"], nom_fichier)
    try:
        df = pd.read_csv(chemin, encoding="utf-8")
        print(f"✓ {nom_fichier} chargé — {df.shape[0]} lignes, {df.shape[1]} colonnes")
        return df
    except FileNotFoundError:
        print(f"✗ Fichier introuvable : {chemin}")
        return None
    except Exception as e:
        print(f"✗ Erreur : {e}")
        return None


def charger_excel(nom_fichier, feuille=0):
    """
    Charge un fichier Excel depuis le dossier data/
    Usage : df = charger_excel("donnees.xlsx", feuille="Sheet1")
    """
    chemin = os.path.join(PATHS["data"], nom_fichier)
    try:
        df = pd.read_excel(chemin, sheet_name=feuille)
        print(f"✓ {nom_fichier} chargé — {df.shape[0]} lignes, {df.shape[1]} colonnes")
        return df
    except FileNotFoundError:
        print(f"✗ Fichier introuvable : {chemin}")
        return None
    except Exception as e:
        print(f"✗ Erreur : {e}")
        return None


def charger_pdf_texte(nom_fichier):
    """
    Extrait le texte d'un PDF depuis le dossier data/
    Usage : texte = charger_pdf_texte("rapport.pdf")
    """
    from pypdf import PdfReader
    chemin = os.path.join(PATHS["data"], nom_fichier)
    try:
        reader = PdfReader(chemin)
        texte = ""
        for page in reader.pages:
            texte += page.extract_text()
        print(f"✓ {nom_fichier} lu — {len(reader.pages)} pages")
        return texte
    except FileNotFoundError:
        print(f"✗ Fichier introuvable : {chemin}")
        return None
    except Exception as e:
        print(f"✗ Erreur : {e}")
        return None


# =============================
# 2. NETTOYAGE DE BASE
# =============================

def nettoyer(df):
    """
    Nettoyage standard :
    - Supprime les doublons
    - Supprime les lignes entièrement vides
    - Normalise les noms de colonnes
    Retourne le df nettoyé + un rapport
    """
    rapport = {}

    rapport["lignes_avant"] = df.shape[0]
    rapport["doublons"]     = df.duplicated().sum()

    df = df.drop_duplicates()
    df = df.dropna(how="all")

    df.columns = (
        df.columns
        .str.strip()
        .str.lower()
        .str.replace(" ", "_")
        .str.replace("é", "e")
        .str.replace("è", "e")
        .str.replace("ê", "e")
        .str.replace("à", "a")
        .str.replace("ô", "o")
    )

    rapport["lignes_apres"] = df.shape[0]
    rapport["colonnes"]     = list(df.columns)

    print(f"✓ Nettoyage terminé")
    print(f"  Lignes avant       : {rapport['lignes_avant']}")
    print(f"  Doublons supprimés : {rapport['doublons']}")
    print(f"  Lignes après       : {rapport['lignes_apres']}")
    print(f"  Colonnes           : {rapport['colonnes']}")

    return df, rapport


# =============================
# 3. RÉSUMÉ STATISTIQUE
# =============================

def resume(df):
    """
    Affiche un résumé statistique rapide du dataframe.
    Usage : resume(df)
    """
    print("\n=== RÉSUMÉ STATISTIQUE ===")
    print(f"Dimensions : {df.shape[0]} lignes × {df.shape[1]} colonnes")
    print(f"\nTypes de colonnes :\n{df.dtypes}")
    print(f"\nValeurs manquantes :\n{df.isnull().sum()}")
    print(f"\nStatistiques :\n{df.describe()}")
    print("==========================\n")


# =============================
# 4. STYLE MATPLOTLIB GLOBAL
# =============================

def style_matplotlib():
    """
    Applique la charte graphique Organon à matplotlib.
    À appeler une fois en début de script.
    """
    plt.rcParams.update({
        "figure.facecolor":     COLORS["white"],
        "axes.facecolor":       COLORS["light"],
        "axes.edgecolor":       "#E5E7EB",
        "axes.labelcolor":      COLORS["dark"],
        "axes.titlesize":       FONTS["title_size"],
        "axes.labelsize":       FONTS["axis_size"],
        "axes.grid":            True,
        "grid.color":           "#E5E7EB",
        "grid.linestyle":       "--",
        "grid.alpha":           0.7,
        "xtick.color":          COLORS["text"],
        "ytick.color":          COLORS["text"],
        "xtick.labelsize":      FONTS["label_size"],
        "ytick.labelsize":      FONTS["label_size"],
        "font.family":          "serif",
        "text.color":           COLORS["dark"],
        "figure.titlesize":     FONTS["title_size"],
    })


# =============================
# 5. APPLIQUER STYLE PLOTLY
# =============================

def appliquer_style(fig, titre):
    """
    Applique la charte graphique Organon à une figure Plotly.
    Usage : fig = appliquer_style(fig, "Mon graphique")
    """
    fig.update_layout(
        **LAYOUT,
        title_text=titre,
        xaxis=dict(
            gridcolor="#E5E7EB",
            linecolor="#E5E7EB",
            tickfont=dict(size=FONTS["label_size"]),
        ),
        yaxis=dict(
            gridcolor="#E5E7EB",
            linecolor="#E5E7EB",
            tickfont=dict(size=FONTS["label_size"]),
        ),
    )

    fig.add_annotation(
        text=ORGANON_SIGNATURE,
        xref="paper", yref="paper",
        x=1, y=-0.12,
        showarrow=False,
        font=dict(
            size=9,
            color=COLORS["primary"],
            family=FONTS["family"],
        ),
        xanchor="right",
    )

    return fig


# =============================
# 6. EXPORTER HTML (Plotly)
# =============================

def exporter(fig, nom_fichier):
    """
    Exporte une figure Plotly en HTML interactif.
    Usage : exporter(fig, "graph1a.html")
    """
    os.makedirs(PATHS["graphiques"], exist_ok=True)
    chemin = os.path.join(PATHS["graphiques"], nom_fichier)
    fig.write_html(
        chemin,
        include_plotlyjs="cdn",
        full_html=True,
    )
    print(f"✓ HTML exporté : {chemin}")


# =============================
# 7. EXPORTER PNG (Matplotlib)
# =============================

def exporter_png(fig_mpl, nom_fichier):
    """
    Exporte une figure Matplotlib en PNG haute résolution.
    Usage : exporter_png(fig, "graph1a.png")
    """
    os.makedirs(PATHS["graphiques"], exist_ok=True)
    chemin = os.path.join(PATHS["graphiques"], nom_fichier)

    # Signature
    fig_mpl.text(
        0.99, 0.01,
        ORGANON_SIGNATURE,
        ha="right", va="bottom",
        fontsize=8,
        color=COLORS["primary"],
        style="italic",
    )

    fig_mpl.savefig(
        chemin,
        dpi=200,
        bbox_inches="tight",
        facecolor=COLORS["white"],
    )
    plt.close(fig_mpl)
    print(f"✓ PNG exporté : {chemin}")


# =============================
# 8. EXPORTER EXCEL
# =============================

def exporter_excel(df, nom_fichier, nom_feuille="Données"):
    """
    Exporte un dataframe en Excel formaté.
    Usage : exporter_excel(df, "resultats.xlsx")
    """
    os.makedirs(PATHS["graphiques"], exist_ok=True)
    chemin = os.path.join(PATHS["graphiques"], nom_fichier)

    with pd.ExcelWriter(chemin, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=nom_feuille, index=False)

        wb = writer.book
        ws = writer.sheets[nom_feuille]

        # Style entête
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="CD7F32")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center")

        # Largeur colonnes auto
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

    print(f"✓ Excel exporté : {chemin}")


# =============================
# 9. EXPORTER PDF
# =============================

def exporter_pdf(nom_fichier, titre, auteur, sections):
    """
    Génère un rapport PDF professionnel.
    sections = liste de dicts :
      {"type": "titre", "texte": "..."}
      {"type": "paragraphe", "texte": "..."}
      {"type": "image", "chemin": "..."}
      {"type": "tableau", "data": [[...]], "entetes": [...]}
      {"type": "saut"}

    Usage :
      exporter_pdf("rapport1.pdf", "Paludisme", "Organon", sections)
    """
    os.makedirs(PATHS["graphiques"], exist_ok=True)
    chemin = os.path.join(PATHS["graphiques"], nom_fichier)

    doc = SimpleDocTemplate(
        chemin,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()

    style_titre = ParagraphStyle(
        "OrganonTitre",
        parent=styles["Heading1"],
        fontSize=18,
        textColor=colors.HexColor(COLORS["dark"]),
        spaceAfter=12,
        fontName="Times-Bold",
    )

    style_sous_titre = ParagraphStyle(
        "OrganonSousTitre",
        parent=styles["Heading2"],
        fontSize=13,
        textColor=colors.HexColor(COLORS["primary"]),
        spaceAfter=8,
        fontName="Times-Bold",
    )

    style_corps = ParagraphStyle(
        "OrganonCorps",
        parent=styles["Normal"],
        fontSize=11,
        textColor=colors.HexColor(COLORS["text"]),
        spaceAfter=8,
        leading=16,
        fontName="Times-Roman",
    )

    style_signature = ParagraphStyle(
        "OrganonSignature",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor(COLORS["primary"]),
        alignment=2,
        fontName="Times-Italic",
    )

    contenu = []

    # Page de garde
    contenu.append(Spacer(1, 3*cm))
    contenu.append(Paragraph(titre, style_titre))
    contenu.append(Spacer(1, 0.5*cm))
    contenu.append(Paragraph(auteur, style_corps))
    contenu.append(Spacer(1, 0.5*cm))
    contenu.append(Paragraph(ORGANON_SIGNATURE, style_signature))
    contenu.append(PageBreak())

    # Sections
    for s in sections:
        if s["type"] == "titre":
            contenu.append(Paragraph(s["texte"], style_sous_titre))

        elif s["type"] == "paragraphe":
            contenu.append(Paragraph(s["texte"], style_corps))

        elif s["type"] == "image":
            contenu.append(Spacer(1, 0.3*cm))
            contenu.append(Image(s["chemin"], width=15*cm, height=9*cm))
            contenu.append(Spacer(1, 0.3*cm))

        elif s["type"] == "tableau":
            data = [s["entetes"]] + s["data"]
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND",  (0,0), (-1,0), colors.HexColor(COLORS["primary"])),
                ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
                ("FONTNAME",    (0,0), (-1,0), "Times-Bold"),
                ("FONTSIZE",    (0,0), (-1,-1), 10),
                ("ROWBACKGROUNDS", (0,1), (-1,-1),
                    [colors.white, colors.HexColor("#F9F7F4")]),
                ("GRID",        (0,0), (-1,-1), 0.5,
                    colors.HexColor("#E5E7EB")),
                ("ALIGN",       (0,0), (-1,-1), "CENTER"),
                ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
                ("PADDING",     (0,0), (-1,-1), 6),
            ]))
            contenu.append(table)
            contenu.append(Spacer(1, 0.4*cm))

        elif s["type"] == "saut":
            contenu.append(PageBreak())

    doc.build(contenu)
    print(f"✓ PDF exporté : {chemin}")